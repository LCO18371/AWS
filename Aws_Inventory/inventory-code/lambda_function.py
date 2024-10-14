import boto3
import openpyxl
from openpyxl import Workbook, load_workbook
from io import BytesIO
from botocore.exceptions import ClientError, NoCredentialsError, PartialCredentialsError, EndpointConnectionError
from datetime import datetime

regions = ['us-east-1', 'us-west-2','us-east-2','ap-south-1']
status = 'Active'
bucket_name = 'task-update-excel-for-all-resource'
account_id = '008893372207'
file_key = 'testing1/AllResource8.xlsx'
s3 = boto3.client('s3')

def update_creation_date():
    global today_date
    if s3_file_exists():
        today_date = datetime.today().strftime('%Y-%m-%d')
    else:
        today_date = 'NA'

def s3_file_exists():
    try:
        s3.head_object(Bucket=bucket_name, Key=file_key)
        return True
    except s3.exceptions.NoSuchKey:
        return False  # The object does not exist
    except ClientError as e:
        print(f"ClientError: {e.response['Error']['Message']}")
        return False
    except (NoCredentialsError, PartialCredentialsError):
        print("Credentials error.")
        return False
    except EndpointConnectionError:
        print("Endpoint connection error.")
        return False


def lambda_handler(event, context):
    global creation_date

    # Update creation_date based on the existence of the file
    update_creation_date()

    try:
        # Check if the file exists in S3
        if today_date == 'NA':
            print(f"{file_key} not found. Creating a new workbook.")
            workbook = Workbook()
            workbook.remove(workbook.active)  # Remove default sheet
        else:
            response = s3.get_object(Bucket=bucket_name, Key=file_key)
            file_content = response['Body'].read()
            workbook = load_workbook(filename=BytesIO(file_content))
            print(f"{file_key} found and loaded.")

        # Resource processing
        process_all_resources(workbook)

        # Save the workbook back to S3
        # Uncomment and implement your save logic here
        # update_excel(workbook, s3)

    except Exception as e:
        print(f"Error: {str(e)}")
        raise e

def process_all_resources(workbook):
    resource_definitions = [
        ('Amplify', get_amplify_apps_info_all_regions, ['Identifier','Name', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
    ]
    '''
    resource_definitions = [
        ('Amplify', get_amplify_apps_info_all_regions, ['Identifier','Name', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('WAF_Web_ACL', get_waf_info_all_regions, ['Identifier','Name', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('Security_Group', get_security_group_info_all_regions, ['Identifier','Name', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('Pinpoint_Application', get_pinpoint_info_all_regions, ['Identifier','Name', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('VPC', get_vpc_info_all_regions, ['Identifier','Name', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('Cognito', get_cognito_info_all_regions, ['Identifier','Name', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),        
        ('Volume', get_target_volume_info_all_regions, ['Identifier','Name', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('Subnet_List', get_subnet_info_all_regions, ['Identifier','Name', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']), 
        ('CodePipeline', get_codepipeline_info_all_regions, ['Identifier', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('EventBridge_Rule', get_eventbridge_info_all_regions, ['Identifier', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('SNS_Topics', get_sns_info_all_region, ['Identifier', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('SQS_Queues', get_sqs_info_all_regions, ['Identifier', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('EC2_Instance', get_ec2_info_all_regions, ['Identifier', 'Name', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('CloudFormation_Stack', get_cloudformation_info_all_regions, ['Identifier', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('Lambda_Function', get_lambda_info_all_regions, ['Identifier', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),        
        ('CloudFront_Distribution', get_cloudfront_info_all_regions, ['Identifier', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('Load_Balancer', get_load_balancer_info_all_regions, ['Identifier', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('Secrets_Manager', get_secrets_manager_info_all_regions, ['Identifier', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('CloudTrail_Trail', get_cloudtrail_info_all_regions, ['Identifier', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('Target_Group', get_target_group_info_all_regions, ['Identifier', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('Config_Service_Rule', get_config_info_all_regions, ['Identifier', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('Dynamodb_Table', get_dynamodb_info_all_regions, ['Identifier', 'Name', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('Step_Function', get_state_machine_info_all_regions, ['Identifier', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('Certificate', get_certificate_info_all_regions, ['Identifier', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('KMS-Key', get_kms_info_all_regions, ['Identifier', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('Directory_Service', get_directory_service_info_all_regions, ['Identifier','Name', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('SES_Verified_Identity', get_ses_info_all_regions, ['Identifier', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('Kinesis_Firehose_Delivery_Stream', get_kinesis_delivery_stream_info_all_regions, ['Identifier', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('WorkGroup', get_athena_info_all_regions, ['Identifier', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('HostedZone', get_route53_info_all_regions, ['Identifier', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('Workspaces', get_workspaces_info_all_regions, ['Identifier', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('Elastic_IP', get_elastic_ip_info_all_regions, ['Identifier','Name', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('EBS_Snapshots', get_ebs_snapshot_info_all_regions, ['Identifier', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('API_Gateway_HTTP', lambda: get_api_gateway_info_all_regions(api_type='HTTP'), ['Identifier','Name', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('API_Gateway_REST', lambda: get_api_gateway_info_all_regions(api_type='REST'), ['Identifier','Name', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
        ('API_Gateway_WEBSOCKET', lambda: get_api_gateway_info_all_regions(api_type='WEBSOCKET'), ['Identifier','Name', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
       
    ]
    '''
    #('Global_Accelerators', get_global_accelerator_info_all_regions, ['Identifier', 'AWS_Service', 'Region', 'Status', 'Creation_Time', 'Deletion_Time', 'Tags']),
    for sheet_name, get_info_func, headers in resource_definitions:
        print(resource_definitions)
        resource_info = get_info_func()
        process_resources(workbook, sheet_name, resource_info, headers)



def update_excel(workbook, s3):
    s3 = boto3.client('s3')
    # Save the workbook to a BytesIO object
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Upload the updated file back to S3
    s3.put_object(Bucket=bucket_name, Key=file_key, Body=output)
    
    print("Excel file updated successfully and uploaded to S3.")

def get_bucket_tags(bucket_name, region):
    s3 = boto3.client('s3', region_name=region)
    try:
        response = s3.get_bucket_tagging(Bucket=bucket_name)
        tags = response['TagSet']
        return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
    except ClientError as e:
        if e.response['Error']['Code'] == 'NoSuchTagSet':
            return 'NA'
        else:
            print(f"Error fetching tags for bucket {bucket_name}: {str(e)}")
            return 'Error fetching tags'

def get_s3_buckets_info_all_regions():
    buckets_info = []
    # Initialize the S3 client
    s3 = boto3.client('s3')
    # Fetch the list of all buckets
    response = s3.list_buckets()
    # Iterate over each bucket in the account
    for bucket in response['Buckets']:
        bucket_name = bucket['Name']
        creation_date = bucket['CreationDate'].strftime('%Y-%m-%d %H:%M:%S')
        try:
            # Retrieve the region for the current bucket
            response = s3.get_bucket_location(Bucket=bucket_name)
            region = response['LocationConstraint']
            # If region is None, it's in us-east-1 (N. Virginia)
            if region is None:
                region = 'us-east-1'
            # Fetch bucket tags
            tags = get_bucket_tags(bucket_name, region)
            # Status placeholder (can be customized based on your needs)

            # Append bucket information to the list
            buckets_info.append({
                'Identifier': bucket_name,
                'AWS_Service': 'S3 Bucket',
                'Region': region,
                'Status': status,
                'Creation_Time': creation_date,
                'Deletion_Time': 'NA',
                'Tags': tags
            })
        except ClientError as e:
            print(f"Error processing bucket {bucket_name}: {str(e)}")
            buckets_info.append({
                'Identifier': bucket_name,
                'AWS_Service': 'S3 Bucket',
                'Region': 'Error fetching tags',
                'Status': status,
                'Creation_Time': 'Error fetching tags',
                'Deletion_Time': 'NA',
                'Tags': 'Error fetching tags'
            })
            continue  # Skip this bucket and continue with the next one
    print(f"Number of S3 buckets: {len(buckets_info)}")
    return buckets_info
    

def get_security_group_info_all_regions():
    security_groups_info = []
    resource_type = 'Security Group'

    for region in regions:
        ec2 = boto3.client('ec2', region_name=region)
        try:
            response = ec2.describe_security_groups()
            print(f"Region: {region} | Number of Security Groups: {len(response['SecurityGroups'])}")
            for sg in response['SecurityGroups']:
                sg_id = sg['GroupId']
                group_name = sg['GroupName']
                vpc_id = sg['VpcId']
                creation_time = 'NA'  # Update this if you can retrieve the actual creation time
                tags = get_security_group_tags(sg_id, region)
                
                response = ec2.describe_tags(Filters=[{'Name': 'resource-id', 'Values': [sg_id]}])
                tags = response.get('Tags', [])
                name = get_tag_value(tags) 
                if not tags:
                    tags = 'NA'
                #print("security_group_name",name) 
                security_groups_info.append({
                    'Identifier': sg_id,
                    'AWS_Service': resource_type,
                    'Name': name,
                    'Region': region,
                    'Status': status,
                    'Creation_Time': creation_time,
                    'Deletion_Time': 'NA',
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching security groups in region {region}: {str(e)}")
            continue  # Skip this region and continue with the next one

    return security_groups_info

def get_security_group_tags(sg_id, region):
    ec2 = boto3.client('ec2', region_name=region)
    try:
        response = ec2.describe_tags(Filters=[{'Name': 'resource-id', 'Values': [sg_id]}])
        tags = response['Tags']
        if not tags:  # If there are no tags
            return 'NA'
        return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
    except ClientError as e:
        print(f"Error fetching tags for Security Group {sg_id}: {str(e)}")
        return 'Error fetching tags'
        
def get_codepipeline_info_all_regions():
    pipelines_info = []
    resource_type = 'CodePipeline'
 
    for region in regions:
        codepipeline = boto3.client('codepipeline', region_name=region)
        try:
            response = codepipeline.list_pipelines()
            pipeline_count = len(response['pipelines'])
            print(f"Region: {region} | Number of Pipelines: {pipeline_count}")
 
            for pipeline in response['pipelines']:
                pipeline_name = pipeline['name']
                try:
                    pipeline_details = codepipeline.get_pipeline(name=pipeline_name)['pipeline']
                    # Note: CodePipeline APIs do not provide direct 'created' info, assuming it's NA
                    creation_date = 'NA'  # Adjust if you have a method to retrieve creation date
                    tags = get_pipeline_tags(pipeline_name, region)
                    pipelines_info.append({
                        'Identifier': pipeline_name,
                        'AWS_Service': resource_type,
                        'Region': region,
                        'Status': status,
                        'Creation_Time': creation_date,
                        'Deletion_Time': 'NA',
                        'Tags': tags
                    })
                except ClientError as e:
                    print(f"Error fetching details for pipeline {pipeline_name} in region {region}: {str(e)}")
                    continue  # Skip this pipeline and continue with the next one
        except ClientError as e:
            print(f"Error fetching pipelines in region {region}: {str(e)}")
            continue  # Skip this region and continue with the next one
 
    return pipelines_info
 
def get_pipeline_tags(pipeline_name, region):
    # Initialize clients for CodePipeline and STS (to get the account ID)
    codepipeline = boto3.client('codepipeline', region_name=region)
    sts = boto3.client('sts')
 
    try:
        # Get the AWS account ID
        account_id = sts.get_caller_identity()["Account"]
        # Print the pipeline name (for debugging purposes)
        # Create the ARN dynamically using the retrieved account ID
        resource_arn = f"arn:aws:codepipeline:{region}:{account_id}:{pipeline_name}"
        # Fetch the tags for the resource
        response = codepipeline.list_tags_for_resource(resourceArn=resource_arn)
        # Get the list of tags, or return 'NA' if no tags are found
        tags = response.get('tags', [])
        if not tags:
            return 'NA'
        # Return the tags in "key=value" format
        return ', '.join([f"{tag['key']}={tag['value']}" for tag in tags])
    except ClientError as e:
        # Handle errors
        print(f"Error fetching tags for pipeline {pipeline_name} in region {region}: {str(e)}")
        return 'Error fetching tags'
        

def get_eventbridge_info_all_regions(): # Add or modify regions as needed
    rules_info = []
    resource_type = 'EventBridge Rule'
 
    for region in regions:
        eventbridge = boto3.client('events', region_name=region)
        rule_count = 0
        # Handle pagination
        next_token = None
        while True:
            if next_token:
                response = eventbridge.list_rules(NextToken=next_token)
            else:
                response = eventbridge.list_rules()
            # Count the number of rules in the current response
            rule_count += len(response['Rules'])
            for rule in response['Rules']:
                rule_name = rule['Name']
                creation_time = rule.get('CreatedBy', 'NA')
 
                # Fetch the tags for the rule
                tags = get_eventbridge_rule_tags(rule_name, region)
 
                rules_info.append({
                    'Identifier': rule_name,
                    'AWS_Service': resource_type,
                    'Region': region,
                    'Status': status,
                    'Creation_Time': creation_time,
                    'Deletion_Time': 'NA',  # You can update this if necessary
                    'Tags': tags
                })
            # Check if there are more rules to fetch
            next_token = response.get('NextToken')
            if not next_token:
                break
        print(f"Region: {region} | Number of Rules: {rule_count}")
 
    return rules_info

def get_amplify_apps_info_all_regions():
    amplify_apps_info = []
    resource_type = 'Amplify App'

    for region in regions:
        amplify = boto3.client('amplify', region_name=region)
        try:
            response = amplify.list_apps()
            print(f"Region: {region} | Number of Amplify Apps: {len(response['apps'])}")
            for app in response.get('apps', []):
                app_id = app['appId']
                app_name = app['name']
                creation_time = app['createdAt'].strftime('%Y-%m-%d %H:%M:%S') if 'createdAt' in app else 'NA'

                amplify_apps_info.append({
                    'Identifier': app_id,
                    'AWS_Service': resource_type,
                    'Name': app_name,
                    'Region': region,
                    'Status': status,
                    'Creation_Time': creation_time,
                    'Deletion_Time': 'NA',
                    'Tags': 'NA'  # Amplify typically doesn't have tags; adjust if needed
                })
        except ClientError as e:
            print(f"Error fetching Amplify apps in region {region}: {str(e)}")
            continue  # Skip this region and continue with the next one

    return amplify_apps_info

def get_eventbridge_rule_tags(rule_name, region):
    eventbridge = boto3.client('events', region_name=region)
    try:
        # Construct the ARN for the EventBridge rule
        account_id = boto3.client('sts').get_caller_identity().get('Account')
        resource_arn = f"arn:aws:events:{region}:{account_id}:rule/{rule_name}"

        response = eventbridge.list_tags_for_resource(ResourceARN=resource_arn)
        tags = response.get('Tags', [])

        # Return "NA" if there are no tags
        if not tags:
            return 'NA'

        # Join tags into a string format
        return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
    except ClientError as e:
        print(f"Error fetching tags for EventBridge rule {rule_name} in region {region}: {str(e)}")
        return 'Error fetching tags'
        
def get_sns_info_all_region():
    topics_info = []
    resource_type = 'SNS Topic'

    for region in regions:
        sns = boto3.client('sns', region_name=region)
        try:
            response = sns.list_topics()
            print(f"Region: {region} | Number of SNS Topics: {len(response['Topics'])}")
            
            for topic in response['Topics']:
                topic_arn = topic['TopicArn']
                topic_name = topic_arn.split(':')[-1]
                tags = get_sns_topic_tags(topic_arn, region)
                topics_info.append({
                    'Identifier': topic_name,
                    'AWS_Service': resource_type,
                    'Region': region,
                    'Status': status,
                    'Creation_Time': 'NA',
                    'Deletion_Time': 'NA',
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching SNS topics in region {region}: {str(e)}")
            continue  # Skip this region and continue with the next one

    return topics_info

def get_sns_topic_tags(topic_arn, region):
    sns = boto3.client('sns', region_name=region)
    try:
        response = sns.list_tags_for_resource(ResourceArn=topic_arn)
        tags = response.get('Tags', [])
        if not tags:  # If there are no tags
            return 'NA'
        return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
    except ClientError as e:
        print(f"Error fetching tags for SNS topic {topic_arn}: {str(e)}")
        return 'Error fetching tags'
        
def get_sqs_info_all_regions():
    queues_info = []
    resource_type = 'SQS Queue'

    for region in regions:
        sqs = boto3.client('sqs', region_name=region)
        try:
            response = sqs.list_queues()
            print(f"Region: {region} | Number of SQS Queues: {len(response.get('QueueUrls', []))}")
            
            for queue_url in response.get('QueueUrls', []):
                queue_name = queue_url.split('/')[-1]
                
                try:
                    queue_attributes = sqs.get_queue_attributes(QueueUrl=queue_url, AttributeNames=['All'])
                    creation_timestamp = queue_attributes['Attributes'].get('CreatedTimestamp', 'NA')
                    
                    # Convert creation timestamp to a readable format if available
                    if creation_timestamp != 'NA':
                        creation_time = datetime.utcfromtimestamp(int(creation_timestamp)).strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        creation_time = 'NA'
                        
                    tags = get_sqs_queue_tags(queue_url, region)
                    
                    queues_info.append({
                        'Identifier': queue_name,
                        'AWS_Service': resource_type,
                        'Region': region,
                        'Status': status,
                        'Creation_Time': creation_time,
                        'Deletion_Time': 'NA',
                        'Tags': tags
                    })
                
                except ClientError as e:
                    print(f"Error fetching attributes for SQS queue {queue_url} in region {region}: {str(e)}")
                    continue  # Skip this queue and continue with the next one
                    
        except ClientError as e:
            print(f"Error fetching SQS queues in region {region}: {str(e)}")
            continue  # Skip this region and continue with the next one

    return queues_info

def get_sqs_queue_tags(queue_url, region):
    sqs = boto3.client('sqs', region_name=region)
    try:
        response = sqs.list_queue_tags(QueueUrl=queue_url)
        
        # Check if 'Tags' key exists in response
        if 'Tags' in response:
            tags_dict = response['Tags']
            if isinstance(tags_dict, dict):
                return ', '.join([f"{key}={value}" for key, value in tags_dict.items()])
            else:
                return 'Unexpected format for tags'
        else:
            return 'No tags found'
        
    except ClientError as e:
        print(f"Error fetching tags for SQS queue {queue_url} in region {region}: {str(e)}")
        return 'Error fetching tags'
        
def get_ec2_info_all_regions():
    ec2_instances_info = []
    resource_type = 'EC2 Instance'

    for region in regions:
        ec2 = boto3.client('ec2', region_name=region)
        try:
            response = ec2.describe_instances()
            print(f"Region: {region} | Number of EC2 Instances: {len(response['Reservations'])}")
            
            for reservation in response['Reservations']:
                for instance in reservation['Instances']:
                    #print("instance", instance)
                    instance_id = instance['InstanceId']
                    # Assuming status as 'Running' for demonstration
                    #status = instance.get('State', {}).get('Name', 'Unknown')
                    launch_time = instance['LaunchTime'].strftime('%Y-%m-%d %H:%M:%S')
                    tags = get_ec2_instance_tags(instance_id, region)
                    name = get_ec2_instance_name(instance_id, region)
                    #print(name)
                    ec2_instances_info.append({
                        'Identifier': instance_id,
                        'Name':name,
                        'AWS_Service': resource_type,
                        'Region': region,
                        'Status': status,
                        'Creation_Time': launch_time,
                        'Deletion_Time': 'NA',
                        'Tags': tags
                    })
        except ClientError as e:
            print(f"Error fetching EC2 instances in region {region}: {str(e)}")
            continue  # Skip this region and continue with the next one

    return ec2_instances_info

def get_ec2_instance_tags(instance_id, region):
    ec2 = boto3.client('ec2', region_name=region)
    try:
        response = ec2.describe_tags(Filters=[{'Name': 'resource-id', 'Values': [instance_id]}])
        tags = response.get('Tags', [])
        if not tags:  # If there are no tags
            return 'NA'
        return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
    except ClientError as e:
        print(f"Error fetching tags for EC2 instance {instance_id}: {str(e)}")
        return 'Error fetching tags'
def get_ec2_instance_name(instance_id, region):
    ec2 = boto3.client('ec2', region_name=region)
    try:
        response = ec2.describe_tags(Filters=[{'Name': 'resource-id', 'Values': [instance_id]}])
        tags = response.get('Tags', [])
        if not tags:  # If there are no tags
            return 'NA'
        
        # Search for the tag with the key 'Name'
        for tag in tags:
            if tag['Key'] == 'Name':
                return tag['Value']
        
        return 'Name tag not found'
    
    except ClientError as e:
        print(f"Error fetching tags for EC2 instance {instance_id}: {str(e)}")
        return 'Error fetching tags'
        
def get_cloudformation_info_all_regions(): # Add or modify regions as needed
    stacks_info = []
    resource_type = 'CloudFormation Stack'
    for region in regions:
        cloudformation = boto3.client('cloudformation', region_name=region)
        stack_count = 0
        # Handle pagination
        next_token = None
        while True:
            try:
                if next_token:
                    response = cloudformation.describe_stacks(NextToken=next_token)
                else:
                    response = cloudformation.describe_stacks()
                # Count the number of stacks in the current response
                stack_count += len(response['Stacks'])
                for stack in response['Stacks']:
                    stack_name = stack['StackName']
                    creation_time = stack['CreationTime'].strftime('%Y-%m-%d %H:%M:%S')
                    deletion_time = stack.get('DeletionTime', 'NA')
                    # Fetch the tags for the stack
                    tags = get_cloudformation_stack_tags(stack_name, region)
                    stacks_info.append({
                        'Identifier': stack_name,
                        'AWS_Service': resource_type,
                        'Region': region,
                        'Status': status,
                        'Creation_Time': creation_time,
                        'Deletion_Time': deletion_time,
                        'Tags': tags
                    })
                # Check if there are more stacks to fetch
                next_token = response.get('NextToken')
                if not next_token:
                    break
            except ClientError as e:
                print(f"Error fetching stacks in region {region}: {str(e)}")
                break  # Exit loop if there's an error fetching stacks in this region
        print(f"Region: {region} | Number of Stacks: {stack_count}")
    return stacks_info

def get_cloudformation_stack_tags(stack_name, region):
    cloudformation = boto3.client('cloudformation', region_name=region)
    try:
        response = cloudformation.describe_stacks(StackName=stack_name)
        tags = response['Stacks'][0].get('Tags', [])
        # Return "NA" if there are no tags
        if not tags:
            return 'NA'
        # Join tags into a string format
        return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
    except ClientError as e:
        print(f"Error fetching tags for CloudFormation stack {stack_name} in region {region}: {str(e)}")
        return 'Error fetching tags'        

def get_ebs_snapshot_info_all_regions(): # Add or modify regions as needed
    snapshots_info = []
    resource_type = 'EBS Snapshot'
 
    for region in regions:
        ec2 = boto3.client('ec2', region_name=region)
        snapshot_count = 0
        # Handle pagination
        next_token = None
        while True:
            try:
                if next_token:
                    response = ec2.describe_snapshots(OwnerIds=['self'], NextToken=next_token)
                else:
                    response = ec2.describe_snapshots(OwnerIds=['self'])
                
                # Count the number of snapshots in the current response
                snapshot_count += len(response['Snapshots'])
                for snapshot in response['Snapshots']:
                    snapshot_id = snapshot['SnapshotId']
                    start_time = snapshot['StartTime'].strftime('%Y-%m-%d %H:%M:%S')
                    tags = get_ebs_snapshot_tags(snapshot_id, region, False)
                    multi_tags = get_ebs_snapshot_tags(snapshot_id, region, True)
                    name = get_tag_value(multi_tags)
                    #print("ebs_name",name) 
                    snapshots_info.append({
                        'Identifier': snapshot_id,
                        'AWS_Service': resource_type,
                        'Region': region,
                        'Status': status,  # EBS Snapshots generally have a status of 'available'
                        'Creation_Time': start_time,
                        'Deletion_Time': 'NA',
                        'Tags': tags
                    })
                
                # Check if there are more snapshots to fetch
                next_token = response.get('NextToken')
                if not next_token:
                    break
            except ClientError as e:
                print(f"Error fetching snapshots in region {region}: {str(e)}")
                break  # Exit loop if there's an error fetching snapshots in this region

        print(f"Region: {region} | Number of Snapshots: {snapshot_count}")

    return snapshots_info

def get_ebs_snapshot_tags(snapshot_id, region,check):
    ec2 = boto3.client('ec2', region_name=region)
    try:
        response = ec2.describe_tags(Filters=[{'Name': 'resource-id', 'Values': [snapshot_id]}])
        tags = response['Tags']
        if check:
            return tags
        if not tags:  # Return "NA" if there are no tags
            return 'NA'
        return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
    except ClientError as e:
        print(f"Error fetching tags for EBS snapshot {snapshot_id} in region {region}: {str(e)}")
        return 'Error fetching tags'

def get_elastic_ip_info_all_regions():
    #regions = ['us-east-1', 'us-east-2', 'ap-southeast-1']
    elastic_ips_info = []
    resource_type = 'Elastic IP'

    for region in regions:
        ec2 = boto3.client('ec2', region_name=region)
        response = ec2.describe_addresses()
        num_elastic_ips = len(response['Addresses'])  # Count the number of Elastic IPs
        print(f"Region: {region} | Number of Elastic IPs: {num_elastic_ips}")

        for address in response['Addresses']:
            allocation_id = address.get('AllocationId', 'NA')
            #status = address.get('AssociationId', 'NA')  # Placeholder for status, adjust as needed
            tags = get_elastic_ip_tags(allocation_id, region,False)
            multi_tags = get_elastic_ip_tags(allocation_id, region,True)
            name = get_tag_value(multi_tags)
            #print("elasticip_name",name) 
            elastic_ips_info.append({
                'Identifier': allocation_id,
                'Name': name,
                'AWS_Service': resource_type,
                'Region': region,
                'Status': status,
                'Creation_Time': 'NA',
                'Deletion_Time': 'NA',
                'Tags': tags
            })

    return elastic_ips_info

def get_elastic_ip_tags(allocation_id, region,check):
    ec2 = boto3.client('ec2', region_name=region)
    try:
        response = ec2.describe_tags(Filters=[{'Name': 'resource-id', 'Values': [allocation_id]}])
        tags = response.get('Tags', [])
        if check:
            return tags
        # Return "NA" if there are no tags
        if not tags:
            return 'NA'
        
        # Join tags into a string format
        return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
    except ClientError as e:
        print(f"Error fetching tags for Elastic IP {allocation_id} in region {region}: {str(e)}")
        return 'Error fetching tags'

def get_lambda_info_all_regions():
    lambda_functions_info = []
    resource_type = 'Lambda Function'

    for region in regions:
        lambda_client = boto3.client('lambda', region_name=region)
        next_marker = None  # Initialize marker for pagination
        total_function_count = 0  # Initialize count for each region

        try:
            while True:
                # List functions with pagination
                if next_marker:
                    response = lambda_client.list_functions(Marker=next_marker)
                else:
                    response = lambda_client.list_functions()

                function_count = len(response['Functions'])
                total_function_count += function_count  # Add to the total count for the region
                
                for function in response['Functions']:
                    function_name = function['FunctionName']
                    creation_date = function['LastModified']
                    tags = get_lambda_tags(function_name, region)
                    
                    lambda_functions_info.append({
                        'Identifier': function_name,
                        'AWS_Service': resource_type,
                        'Region': region,
                        'Status': status,  # Placeholder for status
                        'Creation_Time': creation_date,
                        'Deletion_Time': 'NA',
                        'Tags': tags
                    })

                # Check if there are more results to fetch (pagination)
                next_marker = response.get('NextMarker')
                if not next_marker:
                    break  # Exit loop if no more pages
            
            # Print summary of Lambda functions in this region
            print(f"Region: {region} | Total Number of Lambda Functions: {total_function_count}")

        except ClientError as e:
            print(f"Error fetching Lambda functions in region {region}: {str(e)}")

    return lambda_functions_info
def get_lambda_tags(function_name, region):
    lambda_client = boto3.client('lambda', region_name=region)
    try:
        account_id = boto3.client('sts').get_caller_identity().get('Account')
        resource_arn = f'arn:aws:lambda:{region}:{account_id}:function:{function_name}'
        response = lambda_client.list_tags(Resource=resource_arn)
        tags = response.get('Tags', {})
 
        if not tags:
            return 'NA'
        return ', '.join([f"{key}={value}" for key, value in tags.items()])
    except ClientError as e:
        print(f"Error fetching tags for Lambda function {function_name} in region {region}: {str(e)}")
        return 'Error fetching tags'
        
def get_api_gateway_info_all_regions(api_type):
    api_gateway_info = []
    
    if api_type == 'HTTP':
        resource_type = "API Gateway"
    elif api_type == 'REST':
        resource_type = "API Gateway REST API"
    elif api_type == 'WEBSOCKET':
        resource_type = "API Gateway"
    else:
        raise ValueError(f"Unsupported API type: {api_type}")

    for region in regions:
        if api_type == 'HTTP' or api_type == 'WEBSOCKET':
            apigateway = boto3.client('apigatewayv2', region_name=region)
            try:
                response = apigateway.get_apis()
                api_count = len(response['Items'])
                print(f"Region: {region} | Number of {api_type} APIs: {api_count}")

                for api in response['Items']:
                    # Filter WebSocket APIs if 'WEBSOCKET' is the type
                    if api_type == 'WEBSOCKET' and api['ProtocolType'] != 'WEBSOCKET':
                        continue
                    if api_type == 'HTTP' and api['ProtocolType'] != 'HTTP':
                        continue

                    api_id = api['ApiId']
                    creation_date = api['CreatedDate'].strftime('%Y-%m-%d %H:%M:%S')
                    tags = get_api_gateway_tags(api_id, region, api_type,False)
                    name = api['Name']
                    print("api_http_websoket",name)
                    api_gateway_info.append({
                        'Identifier': api_id,
                        'Name':name,
                        'AWS_Service': resource_type,
                        'Region': region,
                        'Status': status,  # Assuming 'status' is defined elsewhere
                        'Creation_Time': creation_date,
                        'Deletion_Time': 'NA',
                        'Tags': tags
                    })
            except ClientError as e:
                print(f"Error fetching {api_type} APIs in region {region}: {str(e)}")
                continue

        elif api_type == 'REST':
            apigateway = boto3.client('apigateway', region_name=region)
            try:
                response = apigateway.get_rest_apis()
                api_count = len(response['items'])
                print(f"Region: {region} | Number of REST APIs: {api_count}")

                for api in response['items']:
                    api_id = api['id']
                    creation_date = api['createdDate'].strftime('%Y-%m-%d %H:%M:%S')
                    tags = get_api_gateway_tags(api_id, region, api_type,False)
                    name = api['name']
                    print("api_rest",name)
                    api_gateway_info.append({
                        'Identifier': api_id,
                        'Name':name,
                        'AWS_Service': resource_type,
                        'Region': region,
                        'Status': status,
                        'Creation_Time': creation_date,
                        'Deletion_Time': 'NA',
                        'Tags': tags
                    })
            except ClientError as e:
                print(f"Error fetching REST APIs in region {region}: {str(e)}")
                continue

        else:
            raise ValueError(f"Unsupported API type: {api_type}")


    return api_gateway_info



def get_api_gateway_tags(api_id, region, api_type,check):
    try:
        if api_type == 'HTTP' or api_type == 'WEBSOCKET':
            apigateway = boto3.client('apigatewayv2', region_name=region)
            response = apigateway.get_api(ApiId=api_id)
        elif api_type == 'REST':
            apigateway = boto3.client('apigateway', region_name=region)
            response = apigateway.get_rest_api(restApiId=api_id)
        else:
            return 'Unsupported API Type'

        tags = response.get('tags' if api_type == 'REST' else 'Tags', {})
        if check:
            return tags
        
        if not tags:
            return 'NA'
        
        return ', '.join([f"{key}={value}" for key, value in tags.items()])
    except ClientError as e:
        print(f"Error fetching tags for API Gateway {api_type} API {api_id} in region {region}: {str(e)}")
        return 'Error fetching tags'



def get_cloudfront_info_all_regions():
    cloudfront = boto3.client('cloudfront')
    distributions_info = []
    resource_type = 'CloudFront Distribution'
    try:
        response = cloudfront.list_distributions()
        distributions = response.get('DistributionList', {}).get('Items', [])
        distribution_count = len(distributions)
        print(f"Number of CloudFront Distributions: {distribution_count}")
 
        # If there are no distributions, return an empty list
        if not distributions:
            return distributions_info  # Return an empty list if no distributions
        for distribution in distributions:
            distribution_id = distribution['Id']
            arn = distribution['ARN']  # Fetch the ARN here
            creation_date = distribution['LastModifiedTime'].strftime('%Y-%m-%d %H:%M:%S')
            #status = distribution.get('Status', 'NA')  # Ensure the correct status is fetched
            tags = get_cloudfront_tags(arn)  # Pass the ARN instead of the distribution ID
            distributions_info.append({
                'Identifier': distribution_id,
                'AWS_Service': resource_type,
                'Region': 'global',  # CloudFront is a global service
                'Status': status,
                'Creation_Time': creation_date,
                'Deletion_Time': 'NA',
                'Tags': tags
            })
    except Exception as e:
        print(f"Error fetching CloudFront distributions: {str(e)}")
 
    return distributions_info 
 
def get_cloudfront_tags(arn):
    cloudfront = boto3.client('cloudfront')
    try:
        response = cloudfront.list_tags_for_resource(Resource=arn)  # ARN is required
        tags = response.get('Tags', {}).get('Items', [])
        # Return "NA" if there are no tags
        if not tags:
            return 'NA'
        # Join tags into a string format
        return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
    except ClientError as e:
        print(f"Error fetching tags for CloudFront distribution {arn}: {str(e)}")
        return 'Error fetching tags'

def get_load_balancer_info_all_regions():
    load_balancers_info = []
    resource_type = 'Load Balancer'

    for region in regions:
        elb = boto3.client('elbv2', region_name=region)
        try:
            response = elb.describe_load_balancers()
            load_balancer_count = len(response['LoadBalancers'])
            print(f"Region: {region} | Number of Load Balancers: {load_balancer_count}")

            for load_balancer in response['LoadBalancers']:
                load_balancer_name = load_balancer['LoadBalancerName']
                creation_date = load_balancer['CreatedTime'].strftime('%Y-%m-%d %H:%M:%S')
                #status = 'NA'  # Placeholder for status, adjust as needed
                tags = get_load_balancer_tags(load_balancer['LoadBalancerArn'], region)
                
                load_balancers_info.append({
                    'Identifier': load_balancer_name,
                    'AWS_Service': resource_type,
                    'Region': region,
                    'Status': status,
                    'Creation_Time': creation_date,
                    'Deletion_Time': 'NA',
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching Load Balancers in region {region}: {str(e)}")
            continue

    return load_balancers_info

def get_load_balancer_tags(load_balancer_arn, region):
    elb = boto3.client('elbv2', region_name=region)
    try:
        response = elb.describe_tags(ResourceArns=[load_balancer_arn])
        tag_descriptions = response.get('TagDescriptions', [])
        
        if tag_descriptions:
            tags = tag_descriptions[0].get('Tags', [])
            if tags:
                return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
            else:
                return 'NA'
        else:
            return 'NA'
    
    except ClientError as e:
        print(f"Error fetching tags for Load Balancer {load_balancer_arn} in region {region}: {str(e)}")
        return 'Error fetching tags'

def get_subnet_info_all_regions():
    subnets_info = []
    resource_type = 'Subnet'

    for region in regions:
        ec2 = boto3.client('ec2', region_name=region)
        try:
            response = ec2.describe_subnets()
            subnet_count = len(response['Subnets'])
            print(f"Region: {region} | Number of Subnets: {subnet_count}")
            for subnet in response['Subnets']:
                subnet_id = subnet['SubnetId']
                #status = 'NA'  # Placeholder for status, adjust as needed
                creation_date = 'NA'  # Placeholder for creation time, adjust if available
                tags = get_subnet_tags(subnet_id, region,False)
                multi_tags = get_subnet_tags(subnet_id, region, True)
                name = get_tag_value(multi_tags)
                #print("subnet_name",name)
                subnets_info.append({
                    'Identifier': subnet_id,
                    'Name': name,
                    'AWS_Service': resource_type,
                    'Region': region,
                    'Status': status,
                    'Creation_Time': creation_date,
                    'Deletion_Time': 'NA',
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching Subnets in region {region}: {str(e)}")
            continue

    return subnets_info

def get_subnet_tags(subnet_id, region,check):
    ec2 = boto3.client('ec2', region_name=region)
    try:
        response = ec2.describe_tags(Filters=[{'Name': 'resource-id', 'Values': [subnet_id]}])
        tags = response.get('Tags', [])
        if check:
            return tags
        if tags:
            return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
        else:
            return 'NA'
    except ClientError as e:
        print(f"Error fetching tags for Subnet {subnet_id} in region {region}: {str(e)}")
        return 'Error fetching tags'

def get_target_volume_info_all_regions():
    target_volumes_info = []
    resource_type = 'Volume'

    for region in regions:
        ec2 = boto3.client('ec2', region_name=region)
        try:
            response = ec2.describe_volumes()
            volume_count = len(response['Volumes'])
            print(f"Region: {region} | Number of Volumes: {volume_count}")

            for volume in response['Volumes']:
                volume_id = volume['VolumeId']
                creation_date = volume['CreateTime'].strftime('%Y-%m-%d %H:%M:%S')
                #status = 'NA'  # Placeholder for status, adjust as needed
                tags = get_target_volume_tags(volume_id, region,False)
                multi_tags = get_target_volume_tags(volume_id, region, True)
                name = get_tag_value(multi_tags)
                #print("ebs_volume_name",name) 

                target_volumes_info.append({
                    'Identifier': volume_id,
                    'Name': name,
                    'AWS_Service': resource_type,
                    'Region': region,
                    'Status': status,
                    'Creation_Time': creation_date,
                    'Deletion_Time': 'NA',
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching Volumes in region {region}: {str(e)}")
            continue

    return target_volumes_info

def get_target_volume_tags(volume_id, region,check):
    ec2 = boto3.client('ec2', region_name=region)
    try:
        response = ec2.describe_tags(Filters=[{'Name': 'resource-id', 'Values': [volume_id]}])
        tags = response.get('Tags', [])
        if check:
            return tags
        if tags:
            return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
        else:
            return 'NA'
    except ClientError as e:
        print(f"Error fetching tags for Volume {volume_id} in region {region}: {str(e)}")
        return 'Error fetching tags'

def get_secrets_manager_info_all_regions():
    secrets_info = []
    resource_type = 'Secrets Manager'

    for region in regions:
        secrets_manager = boto3.client('secretsmanager', region_name=region)
        try:
            response = secrets_manager.list_secrets(MaxResults=100)  # Adjust MaxResults as needed
            secret_count = len(response['SecretList'])
            print(f"Region: {region} | Number of Secrets: {secret_count}")

            for secret in response['SecretList']:
                secret_name = secret['Name']
                secret_id = secret['ARN']
                creation_date = secret['LastChangedDate'].strftime('%Y-%m-%d %H:%M:%S')
                tags = get_secrets_manager_tags(secret.get('Tags', []), region)

                secrets_info.append({
                    'Identifier': secret_id,
                    'AWS_Service': resource_type,
                    'Region': region,
                    'Status': status,  # Placeholder for status, adjust as needed
                    'Creation_Time': creation_date,
                    'Deletion_Time': 'NA',
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching Secrets Manager secrets in {region}: {str(e)}")

    return secrets_info

def get_secrets_manager_tags(tags, region):
    try:
        if tags:
            return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
        else:
            return 'NA'
    except Exception as e:
        print(f"Error processing tags in region {region}: {str(e)}")
        return 'Error fetching tags'

def get_cognito_info_all_regions():
    cognito_info = []
    resource_type_user_pool = 'Cognito User Pool'
    resource_type_identity_pool = 'Cognito Identity Pool'

    for region in regions:
        # Fetch Cognito User Pools
        cognito_idp = boto3.client('cognito-idp', region_name=region)
        try:
            response_user_pools = cognito_idp.list_user_pools(MaxResults=60)
            user_pool_count = len(response_user_pools['UserPools'])
            print(f"Region: {region} | Number of Cognito User Pools: {user_pool_count}")

            for user_pool in response_user_pools['UserPools']:
                user_pool_id = user_pool['Id']
                user_pool_name = user_pool['Name']
                creation_date = user_pool['CreationDate'].strftime('%Y-%m-%d %H:%M:%S')
                tags = get_cognito_tags(user_pool_id, region, account_id)

                cognito_info.append({
                    'Identifier': user_pool_id,
                    'Name': user_pool_name,
                    'AWS_Service': resource_type_user_pool,
                    'Region': region,
                    'Status': status,
                    'Creation_Time': creation_date,
                    'Deletion_Time': 'NA',
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching Cognito User Pools in {region}: {str(e)}")

        # Fetch Cognito Identity Pools
        cognito_identity = boto3.client('cognito-identity', region_name=region)
        try:
            response_identity_pools = cognito_identity.list_identity_pools(MaxResults=60)
            identity_pool_count = len(response_identity_pools['IdentityPools'])
            print(f"Region: {region} | Number of Cognito Identity Pools: {identity_pool_count}")

            for identity_pool in response_identity_pools['IdentityPools']:
                identity_pool_id = identity_pool['IdentityPoolId']
                identity_pool_name = identity_pool['IdentityPoolName']
                # Add the identity pool creation date and other metadata if needed (if available)
                tags = get_identity_pool_tags(identity_pool_id, region, account_id)

                cognito_info.append({
                    'Identifier': identity_pool_id,
                    'Name': identity_pool_name,
                    'AWS_Service': resource_type_identity_pool,
                    'Region': region,
                    'Status': status,  # Placeholder for identity pool status
                    'Creation_Time': 'NA',  # Identity pools don't have creation dates in API responses
                    'Deletion_Time': 'NA',
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching Cognito Identity Pools in {region}: {str(e)}")

    return cognito_info

def get_cognito_tags(user_pool_id, region, account_id):
    # Retrieve user pool tags
    cognito_idp = boto3.client('cognito-idp', region_name=region)
    print(f'arn:aws:cognito-idp:{region}:{account_id}:userpool/{user_pool_id}')
    try:
        response = cognito_idp.list_tags_for_resource(
            ResourceArn=f'arn:aws:cognito-idp:{region}:{account_id}:userpool/{user_pool_id}'
        )
        return response.get('Tags', {})
    except ClientError as e:
        print(f"Error fetching tags for User Pool {user_pool_id}: {str(e)}")
        return {}

def get_identity_pool_tags(identity_pool_id, region, account_id):
    # Retrieve identity pool tags
    cognito_identity = boto3.client('cognito-identity', region_name=region)
    try:
        response = cognito_identity.list_tags_for_resource(
            ResourceArn=f'arn:aws:cognito-identity:{region}:{account_id}:identitypool/{identity_pool_id}'
        )
        return response.get('Tags', {})
    except ClientError as e:
        print(f"Error fetching tags for Identity Pool {identity_pool_id}: {str(e)}")
        return {}
'''
def get_cognito_tags(user_pool_id, region, account_id):
    cognito = boto3.client('cognito-idp', region_name=region)
    try:
        arn = f'arn:aws:cognito-idp:{region}:{account_id}:userpool/{user_pool_id}'
        response = cognito.list_tags_for_resource(ResourceArn=arn)
        tags = response.get('Tags', {})  # Defaults to an empty dictionary if 'Tags' key is not present

        if tags:
            return ', '.join([f"{key}={value}" for key, value in tags.items()])
        else:
            return 'NA'  # Return 'NA' if no tags are found
    except ClientError as e:
        print(f"Error fetching tags for Cognito User Pool {user_pool_id}: {str(e)}")
        return 'Error fetching tags'
'''
def get_cloudtrail_info_all_regions():
    cloudtrail_info = []
    resource_type = 'CloudTrail'

    for region in regions:
        cloudtrail = boto3.client('cloudtrail', region_name=region)
        try:
            response = cloudtrail.describe_trails()
            trail_count = len(response['trailList'])
            print(f"Region: {region} | Number of CloudTrails: {trail_count}")

            for trail in response['trailList']:
                trail_name = trail['Name']
                response_metadata_date = response['ResponseMetadata']['HTTPHeaders']['date']
                trail_creation_time = datetime.strptime(response_metadata_date, '%a, %d %b %Y %H:%M:%S GMT').strftime('%Y-%m-%d %H:%M:%S')
                tags = get_cloudtrail_tags(trail['TrailARN'], region)

                cloudtrail_info.append({
                    'Identifier': trail_name,
                    'AWS_Service': resource_type,
                    'Region': region,
                    'Status': status,  # Placeholder for status, adjust as needed
                    'Creation_Time': trail_creation_time,
                    'Deletion_Time': 'NA',
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching CloudTrails in {region}: {str(e)}")

    return cloudtrail_info

def get_cloudtrail_tags(trail_arn, region):
    cloudtrail = boto3.client('cloudtrail', region_name=region)
    try:
        response = cloudtrail.list_tags(ResourceIdList=[trail_arn])
        tags = response.get('ResourceTagList', [{}])[0].get('TagsList', [])
        
        if tags:
            return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
        else:
            return 'NA'  # Return 'NA' if no tags are found
    except ClientError as e:
        print(f"Error fetching tags for CloudTrail {trail_arn}: {str(e)}")
        return 'Error fetching tags'

def get_target_group_info_all_regions():
    target_group_info = []
    resource_type = 'Target Group'

    for region in regions:
        elbv2 = boto3.client('elbv2', region_name=region)
        try:
            response = elbv2.describe_target_groups()
            target_group_count = len(response['TargetGroups'])
            print(f"Region: {region} | Number of Target Groups: {target_group_count}")

            for target_group in response['TargetGroups']:
                target_group_arn = target_group['TargetGroupArn']
                target_group_name = target_group['TargetGroupName']
                creation_date = "NA"  # Creation date is not directly available from describe_target_groups
                tags = get_target_group_tags(target_group_arn, region)
                
                target_group_info.append({
                    'Identifier': target_group_arn,
                    'AWS_Service': resource_type,
                    'Region': region,
                    'Status': status,  # Placeholder for status, adjust as needed
                    'Creation_Time': creation_date,
                    'Deletion_Time': 'NA',
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching Target Groups in {region}: {str(e)}")

    return target_group_info

def get_target_group_tags(target_group_arn, region):
    elbv2 = boto3.client('elbv2', region_name=region)
    try:
        response = elbv2.describe_tags(ResourceArns=[target_group_arn])
        tags = response.get('TagDescriptions', [{}])[0].get('Tags', [])
        
        if tags:
            return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
        else:
            return 'NA'  # Return 'NA' if no tags are found
    except ClientError as e:
        print(f"Error fetching tags for Target Group {target_group_arn}: {str(e)}")
        return 'Error fetching tags'
        
def get_config_info_all_regions():
    config_info = []
    resource_type = 'Config Rule'

    for region in regions:
        config = boto3.client('config', region_name=region)
        try:
            response = config.describe_config_rules()
            config_rule_count = len(response['ConfigRules'])
            print(f"Region: {region} | Number of Config Rules: {config_rule_count}")

            for config_rule in response['ConfigRules']:
                config_rule_name = config_rule['ConfigRuleName']
                config_rule_id = config_rule['ConfigRuleId']
                creation_date = 'NA'  # AWS Config rules don't have a direct creation date

                tags = get_config_tags(config_rule_id, region)

                config_info.append({
                    'Identifier': config_rule_name,
                    'AWS_Service': resource_type,
                    'Region': region,
                    'Status': status,
                    'Creation_Time': creation_date,
                    'Deletion_Time': 'NA',
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching Config rules in region {region}: {str(e)}")
            continue

    return config_info

def get_config_tags(config_rule_id, region):
    config = boto3.client('config', region_name=region)
    try:
        response = config.list_tags_for_resource(ResourceArn=f'arn:aws:config:{region}:{account_id}:config-rule/{config_rule_id}')
        tags = response.get('Tags', [])
        if tags:
            return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
        else:
            return 'NA'
    except ClientError as e:
        print(f"Error fetching tags for Config rule {config_rule_id} in region {region}: {str(e)}")
        return 'Error fetching tags'

def get_dynamodb_info_all_regions():
    dynamodb_info = []
    resource_type = 'DynamoDB Table'

    # Add or modify regions as needed
    regions = ['us-east-1', 'us-east-2', 'ap-southeast-1']
    
    for region in regions:
        dynamodb = boto3.client('dynamodb', region_name=region)
        table_count = 0
        last_evaluated_table_name = None

        while True:
            try:
                # Handle pagination with LastEvaluatedTableName
                if last_evaluated_table_name:
                    response = dynamodb.list_tables(ExclusiveStartTableName=last_evaluated_table_name)
                else:
                    response = dynamodb.list_tables()

                table_names = response['TableNames']
                table_count += len(table_names)
                
                for table_name in table_names:
                    # Get table description for additional details
                    table_description = dynamodb.describe_table(TableName=table_name)['Table']
                    table_arn = table_description['TableArn']
                    creation_date = table_description['CreationDateTime'].strftime('%Y-%m-%d %H:%M:%S')
                    status = table_description['TableStatus']
                    
                    # Fetch the tags for the table
                    tags = get_dynamodb_tags(table_arn, region)

                    # Append the table info to the list
                    dynamodb_info.append({
                        'Identifier': table_name,
                        'AWS_Service': resource_type,
                        'Region': region,
                        'Status': status,
                        'Creation_Time': creation_date,
                        'Deletion_Time': 'NA',
                        'Tags': tags
                    })

                # Check if there are more tables to fetch
                last_evaluated_table_name = response.get('LastEvaluatedTableName')
                if not last_evaluated_table_name:
                    break

            except ClientError as e:
                print(f"Error fetching DynamoDB tables in region {region}: {str(e)}")
                break  # Exit loop if there's an error fetching tables in this region

        print(f"Region: {region} | Number of DynamoDB Tables: {table_count}")

    return dynamodb_info

def get_dynamodb_tags(table_arn, region):
    dynamodb = boto3.client('dynamodb', region_name=region)
    try:
        response = dynamodb.list_tags_of_resource(ResourceArn=table_arn)
        tags = response.get('Tags', [])
        if tags:
            return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
        else:
            return 'NA'
    except ClientError as e:
        print(f"Error fetching tags for DynamoDB table {table_arn} in region {region}: {str(e)}")
        return 'Error fetching tags'

def get_state_machine_info_all_regions():
    state_machine_info = []
    resource_type = 'step function'

    for region in regions:
        sfn = boto3.client('stepfunctions', region_name=region)
        try:
            response = sfn.list_state_machines()
            state_machine_count = len(response['stateMachines'])
            print(f"Region: {region} | Number of State Machines: {state_machine_count}")

            for state_machine in response['stateMachines']:
                state_machine_arn = state_machine['stateMachineArn']
                creation_date = 'NA'  # Step Functions don't have a direct creation date in the list response
                tags = get_state_machine_tags(state_machine_arn, region)

                state_machine_info.append({
                    'Identifier': state_machine_arn,
                    'AWS_Service': resource_type,
                    'Region': region,
                    'Status': status,
                    'Creation_Time': creation_date,
                    'Deletion_Time': 'NA',
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching State Machines in {region}: {str(e)}")
            continue

    return state_machine_info

def get_state_machine_tags(state_machine_arn, region):
    sfn = boto3.client('stepfunctions', region_name=region)
    try:
        response = sfn.list_tags_for_resource(resourceArn=state_machine_arn)
        tags = response.get('tags', [])
        if tags:
            return ', '.join([f"{tag['key']}={tag['value']}" for tag in tags])
        else:
            return 'NA'
    except ClientError as e:
        print(f"Error fetching tags for State Machine {state_machine_arn}: {str(e)}")
        return 'Error fetching tags'

def get_certificate_info_all_regions():
    certificate_info = []
    resource_type = 'ACM'

    for region in regions:
        acm = boto3.client('acm', region_name=region)
        try:
            response = acm.list_certificates()
            cert_count = len(response['CertificateSummaryList'])
            print(f"Region: {region} | Number of Certificates: {cert_count}")

            for certificate in response['CertificateSummaryList']:
                certificate_arn = certificate['CertificateArn']
                cert_details = acm.describe_certificate(CertificateArn=certificate_arn)['Certificate']
                creation_date = cert_details['CreatedAt'].strftime('%Y-%m-%d %H:%M:%S')
                tags = get_certificate_tags(certificate_arn, region)

                certificate_info.append({
                    'Identifier': certificate_arn,
                    'AWS_Service': resource_type,
                    'Region': region,
                    'Status': status,
                    'Creation_Time': creation_date,
                    'Deletion_Time': 'NA',
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching Certificates in {region}: {str(e)}")
            continue

    return certificate_info

def get_certificate_tags(certificate_arn, region):
    acm = boto3.client('acm', region_name=region)
    try:
        response = acm.list_tags_for_certificate(CertificateArn=certificate_arn)
        tags = response.get('Tags', [])
        if tags:
            return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
        else:
            return 'NA'
    except ClientError as e:
        print(f"Error fetching tags for Certificate {certificate_arn}: {str(e)}")
        return 'Error fetching tags'

def get_kms_info_all_regions():
    kms_info = []
    resource_type = 'KMS Key'

    for region in regions:
        kms = boto3.client('kms', region_name=region)
        try:
            response = kms.list_keys()
            key_count = len(response['Keys'])
            print(f"Region: {region} | Number of KMS Keys: {key_count}")

            for key in response['Keys']:
                key_id = key['KeyId']
                key_metadata = kms.describe_key(KeyId=key_id)['KeyMetadata']
                creation_date = key_metadata['CreationDate'].strftime('%Y-%m-%d %H:%M:%S')
                tags = get_kms_tags(key_id, region)

                kms_info.append({
                    'Identifier': key_id,
                    'AWS_Service': resource_type,
                    'Region': region,
                    'Status': status,
                    'Creation_Time': creation_date,
                    'Deletion_Time': 'NA',
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching KMS keys in {region}: {str(e)}")
            continue

    return kms_info

def get_kms_tags(key_id, region):
    kms = boto3.client('kms', region_name=region)
    try:
        response = kms.list_resource_tags(KeyId=key_id)
        tags = response.get('Tags', [])
        if tags:
            return ', '.join([f"{tag['TagKey']}={tag['TagValue']}" for tag in tags])
        else:
            return 'NA'
    except ClientError as e:
        print(f"Error fetching tags for KMS key {key_id}: {str(e)}")
        return 'Error fetching tags'

def get_directory_service_info_all_regions():
    directory_info = []
    resource_type = 'Directory'

    for region in regions:
        ds = boto3.client('ds', region_name=region)
        try:
            response = ds.describe_directories()
            directory_count = len(response['DirectoryDescriptions'])
            print(f"Region: {region} | Number of Directories: {directory_count}")

            for directory in response['DirectoryDescriptions']:
                directory_id = directory['DirectoryId']
                creation_date = directory['LaunchTime'].strftime('%Y-%m-%d %H:%M:%S')
                tags = get_directory_service_tags(directory_id, region)
                name = directory['Name']
                directory_info.append({
                    'Identifier': directory_id,
                    'Name': name,
                    'AWS_Service': resource_type,
                    'Region': region,
                    'Status': status,
                    'Creation_Time': creation_date,
                    'Deletion_Time': 'NA',
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching Directories in {region}: {str(e)}")
            continue

    return directory_info

def get_directory_service_tags(directory_id, region):
    ds = boto3.client('ds', region_name=region)
    try:
        response = ds.list_tags_for_resource(ResourceId=directory_id)
        tags = response.get('Tags', [])
        if tags:
            return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
        else:
            return 'NA'
    except ClientError as e:
        print(f"Error fetching tags for Directory {directory_id}: {str(e)}")
        return 'Error fetching tags'

def get_ses_info_all_regions():
    ses_info = []
    resource_type = 'SES Identity'

    for region in regions:
        ses = boto3.client('ses', region_name=region)
        try:
            response = ses.list_identities()
            identity_count = len(response['Identities'])
            print(f"Number of SES identities in {region}: {identity_count}")
            for identity in response['Identities']:
                print(identity)
                creation_date = 'NA'
                tags = get_ses_tags(identity, region)
                ses_info.append({
                    'Identifier': identity,
                    'AWS_Service': resource_type,
                    'Region': region,
                    'Status': status,
                    'Creation_Time': creation_date,
                    'Deletion_Time': 'NA',
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching SES identities in {region}: {str(e)}")
            continue

    return ses_info

def get_ses_tags(identity, region):
    ses = boto3.client('sesv2', region_name=region)
    try:
        response = ses.list_tags_for_resource(ResourceArn=f'arn:aws:ses:{region}:{account_id}:identity/{identity}')
        tags = response.get('Tags', [])
        if not tags:
            return 'NA'
        return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
    except ClientError as e:
        print(f"Error fetching tags for SES identity {identity}: {str(e)}")
        return 'Error fetching tags'

def get_kinesis_delivery_stream_info_all_regions():
    delivery_stream_info = []
    resource_type = 'Delivery Stream'

    for region in regions:
        firehose = boto3.client('firehose', region_name=region)
        try:
            response = firehose.list_delivery_streams()
            stream_count = len(response['DeliveryStreamNames'])
            print(f"Number of Kinesis Delivery Streams in {region}: {stream_count}")
            for delivery_stream_name in response['DeliveryStreamNames']:
                delevery_stream_temp = firehose.describe_delivery_stream(DeliveryStreamName=delivery_stream_name)['DeliveryStreamDescription']
                creation_date = delevery_stream_temp['CreateTimestamp'].strftime('%Y-%m-%d %H:%M:%S')
                status = delevery_stream_temp.get('DeliveryStreamStatus', 'NA')
                tags = get_kinesis_delivery_stream_tags(delivery_stream_name, region)
                delivery_stream_info.append({
                    'Identifier': delivery_stream_name,
                    'AWS_Service': resource_type,
                    'Region': region,
                    'Status': status,
                    'Creation_Time': creation_date,
                    'Deletion_Time': 'NA',
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching Delivery Streams in {region}: {str(e)}")
            continue

    return delivery_stream_info

def get_kinesis_delivery_stream_tags(delivery_stream_arn, region):
    firehose = boto3.client('firehose', region_name=region)
    try:
        response = firehose.list_tags_for_delivery_stream(DeliveryStreamName=delivery_stream_arn)
        tags = response.get('Tags', [])
        if not tags:
            return 'NA'
        return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
    except ClientError as e:
        print(f"Error fetching tags for Delivery Stream {delivery_stream_arn}: {str(e)}")
        return 'Error fetching tags'

def get_waf_info_all_regions():
    waf_info = []
    resource_type = 'WAF'

    for region in regions:
        waf = boto3.client('wafv2', region_name=region)
        try:
            response = waf.list_web_acls(Scope='REGIONAL')
            acl_count = len(response['WebACLs'])
            print(f"Number of WAF Web ACLs in {region}: {acl_count}")
            for web_acl in response['WebACLs']:
                web_acl_id = web_acl['Id']
                name = web_acl['Name']
                creation_date = 'NA'
                status = 'Active'
                tags = 'NA'
                waf_info.append({
                    'Identifier': web_acl_id,
                    'Name': name,
                    'AWS_Service': resource_type,
                    'Region': region,
                    'Status': status,
                    'Creation_Time': creation_date,
                    'Deletion_Time': 'NA',
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching WAF web ACLs in {region}: {str(e)}")

    return waf_info

def get_athena_info_all_regions():
    athena_info = []
    resource_type = 'Athena Workgroup'

    for region in regions:
        athena = boto3.client('athena', region_name=region)
        try:
            response = athena.list_work_groups()
            workgroup_count = len(response['WorkGroups'])
            print(f"Number of Athena Workgroups in {region}: {workgroup_count}")
            for workgroup in response['WorkGroups']:
                workgroup_name = workgroup['Name']
                creation_date = workgroup['CreationTime'].strftime('%Y-%m-%d %H:%M:%S')
                tags = get_athena_tags(workgroup_name, region)
                athena_info.append({
                    'Identifier': workgroup_name,
                    'AWS_Service': resource_type,
                    'Region': region,
                    'Status': status,
                    'Creation_Time': creation_date,
                    'Deletion_Time': 'NA',
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching Athena workgroups in {region}: {str(e)}")
            continue

    return athena_info

def get_athena_tags(workgroup_arn, region):
    athena = boto3.client('athena', region_name=region)
    try:
        response = athena.list_tags_for_resource(ResourceARN=f"arn:aws:athena:{region}:{account_id}:workgroup/{workgroup_arn}")
        tags = response.get('Tags', [])
        if not tags:
            return 'NA'
        return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
    except ClientError as e:
        print(f"Error fetching tags for Athena workgroup {workgroup_arn}: {str(e)}")
        return 'Error fetching tags'

# Remaining functions for Workspaces, Route 53, and Global Accelerator follow the same pattern as above...

def get_vpc_info_all_regions():
    vpc_info = []
    resource_type = 'VPC'

    for region in regions:
        ec2 = boto3.client('ec2', region_name=region)
        try:
            # Fetch VPCs for the given region
            response = ec2.describe_vpcs()
            num_vpcs = len(response['Vpcs'])  # Count the number of VPCs in the region
            print(f"Region: {region} | Number of VPCs: {num_vpcs}")

            for vpc in response['Vpcs']:
                vpc_id = vpc['VpcId']
                creation_date = 'NA'  # AWS API doesn't provide creation date for VPCs
                tags = get_vpc_tags(vpc_id, region,False)
                mult_tags = get_vpc_tags(vpc_id, region,True)
                name = get_tag_value(mult_tags)
                vpc_info.append({
                    'Identifier': vpc_id,
                    'Name': name,
                    'AWS_Service': resource_type,
                    'Region': region,
                    'Status': status,  # Assuming 'status' is a global variable or defined elsewhere
                    'Creation_Time': creation_date,
                    'Deletion_Time': 'NA',  # Placeholder for deletion time
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching VPCs in {region}: {str(e)}")
            continue  # Skip this region and move to the next

    return vpc_info

def get_vpc_tags(vpc_id, region,check):
    ec2 = boto3.client('ec2', region_name=region)
    try:
        # Fetch tags for the VPC
        response = ec2.describe_tags(Filters=[{'Name': 'resource-id', 'Values': [vpc_id]}])
        tags = response.get('Tags', [])
        if check:
            return tags
        if tags:
            return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
        else:
            return 'NA'  # No tags found, return 'NA'
    except ClientError as e:
        print(f"Error fetching tags for VPC {vpc_id} in {region}: {str(e)}")
        return 'Error fetching tags'
    
def get_pinpoint_info_all_regions():
    pinpoint_info = []
    resource_type = 'Pinpoint'

    for region in regions:
        pinpoint = boto3.client('pinpoint', region_name=region)
        try:
            # Fetch all Pinpoint apps in the region
            response = pinpoint.get_apps()
            num_apps = len(response['ApplicationsResponse']['Item'])
            print(f"Region: {region} | Number of Pinpoint Apps: {num_apps}")

            for app in response['ApplicationsResponse']['Item']:
                app_id = app['Id']
                app_name = app['Name']
                creation_time = datetime.strptime(app['CreationDate'], '%Y-%m-%dT%H:%M:%S.%fZ').strftime('%Y-%m-%d %H:%M:%S')

                tags = get_pinpoint_tags(app_id, region)

                pinpoint_info.append({
                    'Identifier': app_id,
                    'Name': app_name,
                    'AWS_Service': resource_type,
                    'Region': region,
                    'Status': status,  # Assuming 'status' is defined elsewhere
                    'Creation_Time': creation_time,
                    'Deletion_Time': 'NA',  # Placeholder, adjust if needed
                    'Tags': tags
                })

        except ClientError as e:
            print(f"Error fetching Pinpoint apps in {region}: {str(e)}")
            continue  # Skip this region and continue with the next

    return pinpoint_info

def get_pinpoint_tags(app_id, region):
    pinpoint = boto3.client('pinpoint', region_name=region)
    try:
        # Fetch tags for the Pinpoint app
        response = pinpoint.list_tags_for_resource(ResourceArn=f'arn:aws:mobiletargeting:{region}:{account_id}:apps/{app_id}')
        tags = response.get('Tags', {})

        if tags:
            return ', '.join([f"{key}={value}" for key, value in tags.items()])
        else:
            return 'NA'  # No tags available

    except ClientError as e:
        print(f"Error fetching tags for Pinpoint app {app_id} in {region}: {str(e)}")
        return 'Error fetching tags'


def get_route53_info_all_regions():
    route53_info = []
    resource_type = 'Route 53 Hosted Zone'

    for region in regions:
        route53 = boto3.client('route53', region_name=region)
        try:
            # Fetch all Route 53 hosted zones
            response = route53.list_hosted_zones()
            num_hosted_zones = len(response['HostedZones'])
            print(f"Region: {region} | Number of Hosted Zones: {num_hosted_zones}")

            for hosted_zone in response['HostedZones']:
                hosted_zone_id = hosted_zone['Id'].split('/')[-1]
                creation_date = 'NA'  # Update this if you can get the creation date

                # Fetch tags for the hosted zone
                tags = get_route53_tags(hosted_zone_id, region)

                route53_info.append({
                    'Identifier': hosted_zone_id,
                    'AWS_Service': resource_type,
                    'Region': region,
                    'Status': status,  # Assuming 'status' is defined elsewhere
                    'Creation_Time': creation_date,
                    'Deletion_Time': 'NA',  # Adjust as needed
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching Route 53 hosted zones in {region}: {str(e)}")
            continue  # Skip this region and move to the next one

    return route53_info

def get_route53_tags(hosted_zone_id, region):
    route53 = boto3.client('route53', region_name=region)
    try:
        # Fetch the tags for the hosted zone
        response = route53.list_tags_for_resource(ResourceType='hostedzone', ResourceId=hosted_zone_id)
        tags = response.get('ResourceTagSet', {}).get('Tags', [])

        if tags:
            return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
        else:
            return 'NA'  # No tags available
    except ClientError as e:
        print(f"Error fetching tags for Route 53 hosted zone {hosted_zone_id}: {str(e)}")
        return 'Error fetching tags'


def get_workspaces_info_all_regions():
    workspaces_info = []
    resource_type = 'Workspace'

    for region in regions:
        try:
            # Initialize the WorkSpaces client for the region
            workspaces = boto3.client('workspaces', region_name=region)
            # Fetch all WorkSpaces in the current region
            response = workspaces.describe_workspaces()
            num_workspaces = len(response['Workspaces'])
            print(f"Region: {region} | Number of WorkSpaces: {num_workspaces}")

            for workspace in response['Workspaces']:
                workspace_id = workspace['WorkspaceId']
                creation_date = 'NA'  # Update this if you can get the creation date
                tags = get_workspaces_tags(workspace_id, region)

                workspaces_info.append({
                    'Identifier': workspace_id,
                    'AWS_Service': resource_type,
                    'Region': region,
                    'Status': status,  # Assuming 'status' is defined elsewhere
                    'Creation_Time': creation_date,
                    'Deletion_Time': 'NA',  # Adjust as needed
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching WorkSpaces in {region}: {str(e)}")
            continue  # Skip this region and continue with the next one
        except EndpointConnectionError as e:
            print(f"Could not connect to WorkSpaces endpoint in {region}: {str(e)}")
            continue  # Skip this region and try the next one

    return workspaces_info


def get_workspaces_tags(workspace_id, region):
    try:
        workspaces = boto3.client('workspaces', region_name=region)
   
        # Fetch tags for the workspace
        response = workspaces.describe_tags(ResourceId=workspace_id)
        tags = response.get('TagList', [])
        
        if tags:
            return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
        else:
            return 'NA'  # No tags available
    except ClientError as e:
        print(f"Error fetching tags for Workspace {workspace_id}: {str(e)}")
        return 'Error fetching tags'
        

def get_global_accelerator_info_all_regions():
    global_accelerator_info = []
    resource_type = 'Global Accelerator'

    for region in regions:
        ga = boto3.client('globalaccelerator', region_name=region)
        try:
            # Fetch all Global Accelerators in the current region
            response = ga.list_accelerators()
            num_accelerators = len(response['Accelerators'])
            print(f"Region: {region} | Number of Global Accelerators: {num_accelerators}")

            for accelerator in response['Accelerators']:
                accelerator_arn = accelerator['AcceleratorArn']
                creation_date = accelerator['CreatedTime'].strftime('%Y-%m-%d %H:%M:%S')
                tags = get_global_accelerator_tags(accelerator_arn, region)

                global_accelerator_info.append({
                    'Identifier': accelerator_arn,
                    'AWS_Service': resource_type,
                    'Region': region,
                    'Status': status,  # Assuming 'status' is defined elsewhere
                    'Creation_Time': creation_date,
                    'Deletion_Time': 'NA',  # Adjust as needed
                    'Tags': tags
                })
        except ClientError as e:
            print(f"Error fetching Global Accelerators in {region}: {str(e)}")
            continue  # Skip this region and continue with the next one

    return global_accelerator_info

def get_global_accelerator_tags(accelerator_arn, region):
    ga = boto3.client('globalaccelerator', region_name=region)
    try:
        # Fetch tags for the Global Accelerator
        response = ga.list_tags_for_resource(ResourceArn=accelerator_arn)
        tags = response.get('Tags', [])

        if tags:
            return ', '.join([f"{tag['Key']}={tag['Value']}" for tag in tags])
        else:
            return 'NA'  # No tags available
    except ClientError as e:
        print(f"Error fetching tags for Global Accelerator {accelerator_arn}: {str(e)}")
        return 'Error fetching tags'

def process_resources(workbook, s3_sheet_name, s3_buckets_info, headers):
    # Select or create the sheet
    if s3_sheet_name in workbook.sheetnames:
        sheet = workbook[s3_sheet_name]
    else:
        sheet = workbook.create_sheet(title=s3_sheet_name)

    # Update headers explicitly, even if the sheet already has data
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header)

    # Get existing data from the sheet
    existing_data = {}
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # Start from row 2
        existing_row_tuple = tuple(row)
        existing_data[existing_row_tuple] = row_index  # Store the row index for later reference

    print("existing_data", existing_data)
    # Set of new rows to track what is in s3_buckets_info
    new_data_set = set()
    # Check if any header contains the word 'name'
    name_header_exists = any('name' in h.lower() for h in headers)
    print("name_header",name_header_exists)
    # Iterate over bucket information
    for bucket_info in s3_buckets_info:
        new_row = []
        for header in headers:
            value = bucket_info.get(header, '')
            # Convert datetime with timezone info to naive datetime
            if isinstance(value, datetime) and value.tzinfo is not None:
                value = value.replace(tzinfo=None)

            # Convert unsupported types to string
            if isinstance(value, (dict, list)):
                value = str(value)

            new_row.append(value)

        # Convert new_row to a tuple for comparison
        new_row_tuple = tuple(new_row)
        new_data_set.add(new_row_tuple)
        print("new_row_tuple",new_row_tuple)
        Identifier_exist_in_excel(existing_data,new_row_tuple, name_header_exists,sheet)
    Deleted_row_updated_in_excel(existing_data,new_data_set,name_header_exists,sheet)
    
    # Call the function to update the workbook in S3
    update_excel(workbook, s3_sheet_name)

def resource_exists(sheet, resource_id):
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        if row[0] == resource_id:  # Use 'Identifier' for checking
            return True
    return False

def get_tag_value(tags):
    """Return the value of the specified tag key."""
    if not tags:
        return 'NA'
    for tag in tags:
        if tag['Key'] == 'Name':
            return tag['Value']
        else:
            print('NA')
            return 'NA'

def Identifier_exist_in_excel(existing_data , new_row_tuple, name_header_exists,sheet ):
    any_id_to_check = new_row_tuple[0]
    exists = any(any_id_to_check == key[0] for key in existing_data.keys())
    print("exist",exists)
    new_row_list = list(new_row_tuple)
    if not exists:
        print("if conditionworked") 
        if today_date != 'NA':
                print("new_row_list",new_row_list)
                if name_header_exists and new_row_list[5] == 'NA':
                    new_row_list[5] = today_date
                elif new_row_list[4] == 'NA':
                    new_row_list[4] = today_date  
        sheet.append(new_row_list)
    else:
        print("else conditionworked") 
        if new_row_tuple not in existing_data:
            print("new_row_tuple not in existing_data")
            #this code is for upding the exact row in excel
            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False)):
                if name_header_exists:
                    if row[0].value == any_id_to_check:  # Assuming the first column has the identifier
                    # Update only the name and tags columns (let's say they are at index 1 and 2)
                        row[1].value = new_row_tuple[1]  # Name
                        row[7].value = new_row_tuple[7]  # Tags
                        break
                else:
                    if row[0].value == any_id_to_check:  # Assuming the first column has the identifier
                        row[6].value = new_row_tuple[7]  # Tags
                        break
    







    '''
    
        # Check if the row already exists in the sheet
        if new_row_tuple not in existing_data:
            #print("sd",today_date)
            # If it doesn't exist, append it with today's date
            new_row_list = list(new_row_tuple)
            if today_date != 'NA':
                print("new_row_list",new_row_list)
                if name_header_exists and new_row_list[5] == 'NA':

                    new_row_list[5] = today_date
            else:
                if new_row_list[4] == 'NA':
                    new_row_list[4] = today_date  # Assuming the 5th column is for the date
                print("new_row_list1 else condition", new_row_list)
            sheet.append(new_row_list)  # Append the new row with updated date
        else:
            print("row exist")
            # If it exists, you can update it if needed
            row_index = existing_data[new_row_tuple]
            existing_row = tuple(sheet[row_index][col].value for col in range(len(headers)))

            # Update if the row has changed
            if existing_row != new_row_tuple:
                for col_index, value in enumerate(new_row_tuple):
                    sheet.cell(row=row_index, column=col_index + 1).value = value
                sheet.cell(row=row_index, column=4).value = today_date  # Update the date column

    '''
def Deleted_row_updated_in_excel(existing_data , new_data_set, name_header_exists,sheet ):
    # Update status of existing rows Delete status
    for existing_row_tuple, row_index in existing_data.items():
        # If the existing row is not in new_data_set, mark it as deleted
        if existing_row_tuple not in new_data_set:
            if name_header_exists:
                sheet.cell(row=row_index, column=5).value = 'deleted'  # Update the Status column
                if sheet.cell(row=row_index, column=7).value == 'NA':
                    print("deleted update if header",existing_row_tuple)
                    sheet.cell(row=row_index, column=7).value = today_date  
            else:
            #print("existig",existing_row_tuple)
            # Update the Status column (assuming it's at index 4)
                sheet.cell(row=row_index, column=4).value = 'deleted'  # Update the Status column
                if sheet.cell(row=row_index, column=6).value == 'NA':
                    print("deleted update else",existing_row_tuple)
                    sheet.cell(row=row_index, column=6).value = today_date
    
    
# Calling main lambda_handler function 
#lambda_handler()